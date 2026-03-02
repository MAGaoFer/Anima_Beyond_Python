"""Importación de personajes desde fichas Excel (.xlsm) de Ánima."""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path

from modelos.personaje import (
    GuerreroMentalista,
    HechiceroMentalista,
    Mago,
    Mentalista,
    Personaje,
    TA_CODIGOS,
    Warlock,
)

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


def normalizar_numero_excel(valor, default=0):
    """Convierte celdas Excel a entero, tolerando texto mixto como '(46)' o 'X'."""
    if valor is None:
        return default
    if isinstance(valor, bool):
        return int(valor)
    if isinstance(valor, (int, float)):
        return int(valor)

    texto = str(valor).strip()
    if not texto or texto.upper() == "X" or texto.upper() == "#N/A":
        return default

    coincidencia = re.search(r"-?\d+", texto)
    if coincidencia:
        try:
            return int(coincidencia.group(0))
        except ValueError:
            return default
    return default


def _normalizar_texto(texto):
    if texto is None:
        return ""
    t = str(texto).strip().lower()
    t = "".join(c for c in unicodedata.normalize("NFD", t) if unicodedata.category(c) != "Mn")
    return t


def mapear_categoria_excel(categoria, datos):
    """Mapea la categoría del Excel a tipo interno de personaje."""
    cat = _normalizar_texto(categoria)

    if "hechicero mentalista" in cat:
        return "Hechicero mentalista"
    if "guerrero mentalista" in cat:
        return "Guerrero mentalista"
    if "warlock" in cat or "paladin" in cat:
        return "Warlock"
    if "mentalista" in cat:
        return "Mentalista"
    if any(token in cat for token in ("mago", "hechicero", "ilusionista", "conjurador")):
        return "Mago"
    if "novel" in cat:
        if datos.get("zeon", 0) > 0 and datos.get("proyeccion_psiquica", 0) > 0:
            return "Hechicero mentalista"
        if datos.get("zeon", 0) > 0:
            return "Warlock"
        if datos.get("proyeccion_psiquica", 0) > 0:
            return "Guerrero mentalista"
        return "Guerrero"

    return "Guerrero"


def _obtener_hoja(wb, nombre):
    return wb[nombre] if nombre in wb.sheetnames else None


def _texto_celda(ws, celda):
    if ws is None:
        return None
    return ws[celda].value


def _obtener_nombre(ws_general, ws_principal, ws_combate, ruta_excel):
    candidatos = [
        _texto_celda(ws_general, "F22"),
        _texto_celda(ws_principal, "K4"),
        _texto_celda(ws_principal, "AD4"),
        _texto_celda(ws_combate, "M4"),
    ]
    for valor in candidatos:
        if isinstance(valor, str):
            nombre = valor.strip()
            if nombre and "nombre" not in _normalizar_texto(nombre):
                return nombre
    return ruta_excel.stem


def _ta_desde_texto(valor, default="FIL"):
    if valor is None:
        return default
    texto = str(valor).upper()
    for codigo in TA_CODIGOS:
        if codigo in texto:
            return codigo
    return default


def _ta_por_nombre_arma(nombre_arma, default="FIL"):
    nombre = _normalizar_texto(nombre_arma)
    if "escudo" in nombre:
        return "CON"
    if "espada" in nombre or "daga" in nombre or "lanza" in nombre or "hacha" in nombre:
        return "FIL"
    return default


def _extraer_armas(ws_combate):
    if ws_combate is None:
        return [], None

    armas = []
    turno_doble_armas = None
    bloques = [27, 34, 41, 48, 57]

    for inicio in bloques:
        nombre_bloque = _texto_celda(ws_combate, f"D{inicio}")
        nombre_izquierda = _texto_celda(ws_combate, f"E{inicio + 1}") or nombre_bloque
        turno_izq = normalizar_numero_excel(_texto_celda(ws_combate, f"H{inicio + 2}"), default=0)
        ataque_izq = normalizar_numero_excel(_texto_celda(ws_combate, f"I{inicio + 2}"), default=0)
        parada_izq = normalizar_numero_excel(_texto_celda(ws_combate, f"J{inicio + 2}"), default=0)
        danio_izq = normalizar_numero_excel(_texto_celda(ws_combate, f"L{inicio + 2}"), default=0)
        tipo_izq = _ta_desde_texto(_texto_celda(ws_combate, f"D{inicio + 4}"))
        entereza_izq = normalizar_numero_excel(_texto_celda(ws_combate, f"E{inicio + 4}"), default=0)
        rotura_izq = normalizar_numero_excel(_texto_celda(ws_combate, f"F{inicio + 4}"), default=0)

        nombre_izquierda_txt = str(nombre_izquierda).strip() if nombre_izquierda is not None else ""
        if (
            nombre_izquierda_txt
            and nombre_izquierda_txt.upper() != "#N/A"
            and any(v > 0 for v in (turno_izq, ataque_izq, parada_izq, danio_izq, entereza_izq, rotura_izq))
        ):
            tipo_izq = _ta_por_nombre_arma(nombre_izquierda_txt, default=tipo_izq)
            arma = {
                "nombre": nombre_izquierda_txt,
                "turno": turno_izq,
                "daño": danio_izq,
                "rotura": rotura_izq,
                "entereza": entereza_izq,
                "tipo_danio": tipo_izq,
                "habilidad_ataque": ataque_izq,
                "habilidad_parada": parada_izq,
                "es_escudo": "escudo" in _normalizar_texto(nombre_izquierda_txt),
            }
            armas.append(arma)

        nombre_derecha = _texto_celda(ws_combate, f"P{inicio + 1}")
        nombre_derecha_txt = str(nombre_derecha).strip() if nombre_derecha is not None else ""
        turno_der = normalizar_numero_excel(_texto_celda(ws_combate, f"S{inicio + 2}"), default=0)
        ataque_der = normalizar_numero_excel(_texto_celda(ws_combate, f"T{inicio + 2}"), default=0)
        parada_der = normalizar_numero_excel(_texto_celda(ws_combate, f"U{inicio + 2}"), default=0)
        danio_der = normalizar_numero_excel(_texto_celda(ws_combate, f"W{inicio + 2}"), default=0)
        tipo_der = _ta_desde_texto(_texto_celda(ws_combate, f"N{inicio + 4}"))
        entereza_der = normalizar_numero_excel(_texto_celda(ws_combate, f"P{inicio + 4}"), default=0)
        rotura_der = normalizar_numero_excel(_texto_celda(ws_combate, f"Q{inicio + 4}"), default=0)

        if (
            nombre_derecha_txt
            and nombre_derecha_txt.upper() != "#N/A"
            and any(v > 0 for v in (turno_der, ataque_der, parada_der, danio_der, entereza_der, rotura_der))
        ):
            tipo_der = _ta_por_nombre_arma(nombre_derecha_txt, default=tipo_der)
            arma = {
                "nombre": nombre_derecha_txt,
                "turno": turno_der,
                "daño": danio_der,
                "rotura": rotura_der,
                "entereza": entereza_der,
                "tipo_danio": tipo_der,
                "habilidad_ataque": ataque_der,
                "habilidad_parada": parada_der,
                "es_escudo": "escudo" in _normalizar_texto(nombre_derecha_txt),
            }
            armas.append(arma)

        if (
            isinstance(nombre_bloque, str)
            and " y " in _normalizar_texto(nombre_bloque)
            and turno_doble_armas is None
            and turno_izq != 0
        ):
            turno_doble_armas = turno_izq

    dedupe = {}
    for arma in armas:
        clave = (
            arma.get("nombre"),
            arma.get("turno"),
            arma.get("daño"),
            arma.get("rotura"),
            arma.get("entereza"),
            arma.get("tipo_danio"),
            arma.get("habilidad_ataque"),
            arma.get("habilidad_parada"),
        )
        dedupe[clave] = arma

    armas_limpias = list(dedupe.values())
    return armas_limpias, turno_doble_armas


def importar_personaje_desde_excel(ruta_excel):
    """Importa un personaje desde una ficha .xlsm y devuelve (personaje, metadatos)."""
    if openpyxl is None:  # pragma: no cover
        raise RuntimeError("Falta la dependencia 'openpyxl'. Instala requirements.txt")

    ruta = Path(ruta_excel)
    wb = openpyxl.load_workbook(ruta, data_only=True, keep_vba=True)

    ws_general = _obtener_hoja(wb, "General")
    ws_principal = _obtener_hoja(wb, "Principal")
    ws_resumen = _obtener_hoja(wb, "Resumen")
    ws_pds = _obtener_hoja(wb, "PDs")
    ws_combate = _obtener_hoja(wb, "Combate")
    ws_misticos = _obtener_hoja(wb, "Místicos")
    ws_psiquicos = _obtener_hoja(wb, "Psíquicos")

    nombre = _obtener_nombre(ws_general, ws_principal, ws_combate, ruta)
    categoria = _texto_celda(ws_pds, "E11") or "Guerrero"

    puntos_vida = normalizar_numero_excel(_texto_celda(ws_principal, "N11"), default=1)
    puntos_cansancio = normalizar_numero_excel(_texto_celda(ws_principal, "N16"), default=0)
    turno = normalizar_numero_excel(_texto_celda(ws_combate, "H21"), default=0)
    habilidad_ataque = normalizar_numero_excel(_texto_celda(ws_combate, "I21"), default=0)
    habilidad_defensa = normalizar_numero_excel(_texto_celda(ws_combate, "J21"), default=0)
    danio = normalizar_numero_excel(_texto_celda(ws_combate, "L21"), default=0)

    resistencia_fisica = normalizar_numero_excel(_texto_celda(ws_resumen, "F16"), default=0)
    resistencia_magica = normalizar_numero_excel(_texto_celda(ws_resumen, "J16"), default=0)
    resistencia_psiquica = normalizar_numero_excel(_texto_celda(ws_resumen, "N16"), default=0)
    resistencia_venenos = normalizar_numero_excel(_texto_celda(ws_resumen, "R16"), default=0)
    resistencia_enfermedades = normalizar_numero_excel(_texto_celda(ws_resumen, "V16"), default=0)

    armaduras_ta = {
        "FIL": normalizar_numero_excel(_texto_celda(ws_combate, "I16"), default=0),
        "CON": normalizar_numero_excel(_texto_celda(ws_combate, "J16"), default=0),
        "PEN": normalizar_numero_excel(_texto_celda(ws_combate, "K16"), default=0),
        "CAL": normalizar_numero_excel(_texto_celda(ws_combate, "L16"), default=0),
        "ELE": normalizar_numero_excel(_texto_celda(ws_combate, "M16"), default=0),
        "FRI": normalizar_numero_excel(_texto_celda(ws_combate, "N16"), default=0),
        "ENE": normalizar_numero_excel(_texto_celda(ws_combate, "O16"), default=0),
    }
    entereza_armadura = max(
        normalizar_numero_excel(_texto_celda(ws_combate, "P12"), default=0),
        normalizar_numero_excel(_texto_celda(ws_combate, "P13"), default=0),
        normalizar_numero_excel(_texto_celda(ws_combate, "P14"), default=0),
        normalizar_numero_excel(_texto_celda(ws_combate, "P15"), default=0),
    )

    puntos_ki = normalizar_numero_excel(_texto_celda(ws_resumen, "I35"), default=0)
    acumulacion_ki = normalizar_numero_excel(_texto_celda(ws_resumen, "I37"), default=0)

    zeon = normalizar_numero_excel(_texto_celda(ws_misticos, "K18"), default=normalizar_numero_excel(_texto_celda(ws_resumen, "V48"), default=0))
    proyeccion_magica = max(
        normalizar_numero_excel(_texto_celda(ws_misticos, "P12"), default=0),
        normalizar_numero_excel(_texto_celda(ws_misticos, "Q12"), default=0),
        normalizar_numero_excel(_texto_celda(ws_resumen, "J48"), default=0),
    )
    act = normalizar_numero_excel(_texto_celda(ws_misticos, "L12"), default=0)

    potencial_psiquico = max(
        normalizar_numero_excel(_texto_celda(ws_psiquicos, "H11"), default=0),
        normalizar_numero_excel(_texto_celda(ws_psiquicos, "H9"), default=0),
        normalizar_numero_excel(_texto_celda(ws_resumen, "J62"), default=0),
    )
    proyeccion_psiquica = max(
        normalizar_numero_excel(_texto_celda(ws_psiquicos, "P12"), default=0),
        normalizar_numero_excel(_texto_celda(ws_psiquicos, "Q12"), default=0),
        normalizar_numero_excel(_texto_celda(ws_resumen, "K64"), default=0),
    )
    cv_libres = normalizar_numero_excel(_texto_celda(ws_psiquicos, "I17"), default=normalizar_numero_excel(_texto_celda(ws_resumen, "H60"), default=0))

    armas, turno_doble_armas = _extraer_armas(ws_combate)
    principal = armas[0] if armas else {}

    datos_tipado = {
        "zeon": zeon,
        "proyeccion_magica": proyeccion_magica,
        "potencial_psiquico": potencial_psiquico,
        "proyeccion_psiquica": proyeccion_psiquica,
    }
    tipo = mapear_categoria_excel(categoria, datos_tipado)

    comunes = {
        "nombre": nombre,
        "puntos_vida": max(1, puntos_vida),
        "puntos_cansancio": max(0, puntos_cansancio),
        "turno": max(0, turno),
        "daño": max(0, danio),
        "armadura": armaduras_ta.get("FIL", 0),
        "resistencia_fisica": max(0, resistencia_fisica),
        "resistencia_enfermedades": max(0, resistencia_enfermedades),
        "resistencia_venenos": max(0, resistencia_venenos),
        "resistencia_magica": max(0, resistencia_magica),
        "resistencia_psiquica": max(0, resistencia_psiquica),
        "es_pj": True,
        "armaduras_ta": armaduras_ta,
        "entereza_armadura": max(0, entereza_armadura),
    }

    if tipo == "Guerrero":
        personaje = Personaje(
            puntos_ki=max(0, puntos_ki),
            habilidad_ataque=max(0, habilidad_ataque),
            habilidad_defensa=max(0, habilidad_defensa),
            arma_nombre=principal.get("nombre"),
            arma_turno=principal.get("turno"),
            arma_danio=principal.get("daño"),
            arma_rotura=principal.get("rotura"),
            arma_entereza=principal.get("entereza"),
            arma_tipo_danio=principal.get("tipo_danio"),
            armas=armas,
            turno_doble_armas=turno_doble_armas,
            **comunes,
        )
    elif tipo == "Warlock":
        personaje = Warlock(
            puntos_ki=max(0, puntos_ki),
            habilidad_ataque=max(0, habilidad_ataque),
            habilidad_defensa=max(0, habilidad_defensa),
            zeon=max(0, zeon),
            proyeccion_magica=max(0, proyeccion_magica),
            arma_nombre=principal.get("nombre"),
            arma_turno=principal.get("turno"),
            arma_danio=principal.get("daño"),
            arma_rotura=principal.get("rotura"),
            arma_entereza=principal.get("entereza"),
            arma_tipo_danio=principal.get("tipo_danio"),
            armas=armas,
            turno_doble_armas=turno_doble_armas,
            **comunes,
        )
    elif tipo == "Mago":
        personaje = Mago(
            puntos_ki=0,
            zeon=max(0, zeon),
            proyeccion_magica=max(0, proyeccion_magica),
            **comunes,
        )
    elif tipo == "Mentalista":
        personaje = Mentalista(
            puntos_ki=0,
            potencial_psiquico=max(0, potencial_psiquico),
            proyeccion_psiquica=max(0, proyeccion_psiquica),
            cv_libres=max(0, cv_libres),
            **comunes,
        )
    elif tipo == "Guerrero mentalista":
        personaje = GuerreroMentalista(
            puntos_ki=max(0, puntos_ki),
            habilidad_ataque=max(0, habilidad_ataque),
            habilidad_defensa=max(0, habilidad_defensa),
            potencial_psiquico=max(0, potencial_psiquico),
            proyeccion_psiquica=max(0, proyeccion_psiquica),
            cv_libres=max(0, cv_libres),
            arma_nombre=principal.get("nombre"),
            arma_turno=principal.get("turno"),
            arma_danio=principal.get("daño"),
            arma_rotura=principal.get("rotura"),
            arma_entereza=principal.get("entereza"),
            arma_tipo_danio=principal.get("tipo_danio"),
            armas=armas,
            turno_doble_armas=turno_doble_armas,
            **comunes,
        )
    else:
        personaje = HechiceroMentalista(
            puntos_ki=max(0, puntos_ki),
            zeon=max(0, zeon),
            proyeccion_magica=max(0, proyeccion_magica),
            potencial_psiquico=max(0, potencial_psiquico),
            proyeccion_psiquica=max(0, proyeccion_psiquica),
            cv_libres=max(0, cv_libres),
            **comunes,
        )

    personaje.act = max(0, act)
    personaje.acumulacion_ki = max(0, acumulacion_ki)

    metadatos = {
        "ruta": str(ruta),
        "categoria_excel": categoria,
        "tipo_importado": tipo,
        "act": act,
        "acumulacion_ki": acumulacion_ki,
        "armas_detectadas": len(armas),
    }
    return personaje, metadatos
