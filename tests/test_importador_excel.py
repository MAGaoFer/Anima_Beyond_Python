from openpyxl import Workbook

from almacenamiento.importador_excel import (
    _extraer_habilidades_secundarias,
    mapear_categoria_excel,
    normalizar_numero_excel,
)


def test_normalizar_numero_excel_admite_formato_texto():
    assert normalizar_numero_excel("(46)") == 46
    assert normalizar_numero_excel("X", default=0) == 0
    assert normalizar_numero_excel(None, default=7) == 7


def test_mapear_categoria_excel_grupos_principales():
    assert mapear_categoria_excel("Guerrero", {}) == "Guerrero"
    assert mapear_categoria_excel("Tao", {}) == "Guerrero"
    assert mapear_categoria_excel("Warlock", {}) == "Warlock"
    assert mapear_categoria_excel("Hechicero", {}) == "Mago"
    assert mapear_categoria_excel("Mentalista (C)", {}) == "Mentalista"
    assert mapear_categoria_excel("Hechicero mentalista", {}) == "Hechicero mentalista"


def test_mapear_categoria_excel_novel_por_datos():
    assert mapear_categoria_excel("Novel", {"zeon": 100, "proyeccion_psiquica": 80}) == "Hechicero mentalista"
    assert mapear_categoria_excel("Novel", {"zeon": 100, "proyeccion_psiquica": 0}) == "Warlock"
    assert mapear_categoria_excel("Novel", {"zeon": 0, "proyeccion_psiquica": 80}) == "Guerrero mentalista"
    assert mapear_categoria_excel("Novel", {"zeon": 0, "proyeccion_psiquica": 0}) == "Guerrero"


def test_extraer_habilidades_secundarias_extrae_valores_y_preserva_guion():
    wb = Workbook()
    ws = wb.active
    ws.title = "Principal"

    ws["L20"] = "Habilidades Secundarias"
    ws["L22"] = "Atléticas"
    ws["M22"] = "Acrobacias"
    ws["Q22"] = 25

    ws["L29"] = "Sociales"
    ws["M30"] = "Intimidar"
    ws["Q30"] = -25

    ws["L39"] = "Intelectuales"
    ws["M45"] = "Medicina"
    ws["Q45"] = "-"

    ws["L52"] = "Vigor"
    ws["M53"] = "P. Fuerza"
    ws["Q53"] = 40

    ws["L55"] = "Subterfugio"
    ws["M61"] = "Venenos"
    ws["Q61"] = "-"

    ws["L62"] = "Creativas"
    ws["M64"] = "Forja"
    ws["Q64"] = 45

    ws["L74"] = "Especial"

    habilidades = _extraer_habilidades_secundarias(ws)

    assert habilidades["Atléticas"]["Acrobacias"] == 25
    assert habilidades["Sociales"]["Intimidar"] == -25
    assert habilidades["Intelectuales"]["Medicina"] == "-"
    assert habilidades["Vigor"]["P. Fuerza"] == 40
    assert habilidades["Subterfugio"]["Venenos"] == "-"
    assert habilidades["Creativas"]["Forja"] == 45
    assert "Especiales" not in habilidades


def test_extraer_habilidades_secundarias_incluye_especiales_si_hay_habilidad():
    wb = Workbook()
    ws = wb.active
    ws.title = "Principal"

    ws["L20"] = "Habilidades Secundarias"
    ws["L74"] = "Especial"
    ws["M75"] = "Navegación astral"
    ws["Q75"] = 15

    habilidades = _extraer_habilidades_secundarias(ws)

    assert habilidades["Especiales"]["Navegación astral"] == 15
